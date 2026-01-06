from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, Count
from django.db import models
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
import csv
import csv
try:
    import openpyxl
except ImportError:
    openpyxl = None

from .models import Product, StockLog, Category, Sale, SaleItem, Purchase, PurchaseItem, Customer, SalesPerson
from .forms import ProductForm, ImportFileForm, CustomUserCreationForm, CustomUserChangeForm, CustomerForm, SalesPersonForm
from django.db.models.functions import TruncDate
from django.contrib.sessions.models import Session
from django.utils import timezone

@login_required
def dashboard(request):
    total_products = Product.objects.count()
    total_stock_value = Product.objects.aggregate(
        total_value=Sum(F('quantity') * F('purchase_price'))
    )['total_value'] or 0
    
    low_stock_products = Product.objects.filter(quantity__lte=F('min_stock_alert'))
    # Recent Transactions (Sales & Purchases)
    recent_sales = Sale.objects.select_related('user').order_by('-date')[:5]
    recent_purchases = Purchase.objects.select_related('user').order_by('-date')[:5]
    
    # Merge and Sort
    transactions = []
    for s in recent_sales:
        transactions.append({
            'type': 'Sale',
            'id': s.id,
            'order_id': s.order_id,
            'date': s.date,
            'user': s.user,
            'amount': s.total_amount,
            'items_count': s.items.count()
        })
    for p in recent_purchases:
        transactions.append({
            'type': 'Purchase',
            'id': p.id,
            'order_id': p.order_id,
            'date': p.date,
            'user': p.user,
            'amount': p.total_amount,
            'items_count': p.items.count()
        })
        
    # Sort by date descending
    transactions.sort(key=lambda x: x['date'], reverse=True)
    recent_transactions = transactions[:10]

    # Active Users (sessions not expired)
    sessions = Session.objects.filter(expire_date__gte=timezone.now())
    uid_list = []
    for s in sessions:
        data = s.get_decoded()
        uid = data.get('_auth_user_id', None)
        if uid:
            uid_list.append(uid)
    active_users_count = User.objects.filter(id__in=uid_list).count()

    context = {
        'total_products': total_products,
        'total_stock_value': total_stock_value,
        'low_stock_count': low_stock_products.count(),
        'recent_transactions': recent_transactions,
        'low_stock_products': low_stock_products,
        'active_users_count': active_users_count,
    }
    return render(request, 'inventory/dashboard.html', context)

@login_required
def product_list(request):
    products = Product.objects.select_related('category').all()
    query = request.GET.get('q')
    if query:
        products = products.filter(name__icontains=query) | products.filter(code__icontains=query)
    
    context = {'products': products}
    return render(request, 'inventory/product_list.html', context)

@login_required
@permission_required('inventory.add_product', raise_exception=True)
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            StockLog.objects.create(
                product=product,
                user=request.user,
                action='ADD',
                quantity_change=product.quantity,
                reason='Initial Creation'
            )
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'inventory/product_form.html', {'form': form, 'title': 'Add Product'})

@login_required
@permission_required('inventory.change_product', raise_exception=True)
def product_update(request, pk):
    product = get_object_or_404(Product, pk=pk)
    old_qty = product.quantity
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            product = form.save()
            # Log stock change if any
            if product.quantity != old_qty:
                change = product.quantity - old_qty
                StockLog.objects.create(
                    product=product,
                    user=request.user,
                    action='ADJUST',
                    quantity_change=change,
                    reason='Update via Form'
                )
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/product_form.html', {'form': form, 'title': 'Edit Product'})

@login_required
@permission_required('inventory.delete_product', raise_exception=True)
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('product_list')
    return render(request, 'inventory/product_confirm_delete.html', {'product': product})

@login_required
def category_list(request):
    categories = Category.objects.annotate(product_count=Count('products')).all()
    return render(request, 'inventory/category_list.html', {'categories': categories})

@login_required
@permission_required('inventory.add_category', raise_exception=True)
def category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        if name:
            Category.objects.create(name=name, description=description)
            return redirect('category_list')
    return render(request, 'inventory/category_form.html')

@login_required
@permission_required('inventory.delete_category', raise_exception=True)
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('category_list')
    return render(request, 'inventory/category_confirm_delete.html', {'category': category})

@login_required
@permission_required('inventory.add_product', raise_exception=True)
def product_import(request):
    if request.method == 'POST':
        form = ImportFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            created_count = 0
            errors = []
            
            try:
                # Identify file type
                is_excel = file.name.endswith(('.xlsx', '.xls'))
                
                rows = []
                if is_excel:
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    # Skip header
                    iter_rows = ws.iter_rows(values_only=True)
                    next(iter_rows, None) 
                    rows = list(iter_rows)
                else:
                    # CSV handling
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.reader(decoded_file)
                    next(reader, None) # Skip header
                    rows = list(reader)

                for row in rows:
                    try:
                        # Expected format: Name, Code, Category, Purchase Price, Selling Price, Qty, Min Stock
                        # Adjust index checking based on format. Assuming standard 7 columns for now.
                        if not row or len(row) < 3: continue 
                        
                        name = row[0]
                        code = str(row[1])
                        category_name = row[2]
                        p_price = row[3] if len(row) > 3 else 0
                        s_price = row[4] if len(row) > 4 else 0
                        qty = row[5] if len(row) > 5 else 0
                        min_alert = row[6] if len(row) > 6 else 10

                        # Get or Create Category
                        category, _ = Category.objects.get_or_create(name=category_name)
                        
                        # Create Product (avoid duplicates by code)
                        if not Product.objects.filter(code=code).exists():
                            product = Product.objects.create(
                                name=name,
                                code=code,
                                category=category,
                                purchase_price=p_price,
                                selling_price=s_price,
                                quantity=qty,
                                min_stock_alert=min_alert
                            )
                            created_count += 1
                            
                            # Log initial stock
                            StockLog.objects.create(
                                product=product,
                                user=request.user,
                                action='ADD',
                                quantity_change=qty,
                                reason='Bulk Import'
                            )
                    except Exception as e:
                        errors.append(f"Error row {row}: {str(e)}")
                
                if created_count > 0:
                    messages.success(request, f"Successfully imported {created_count} products.")
                else:
                    messages.warning(request, "No new products imported. Check codes for duplicates.")
                    
                if errors:
                    for err in errors[:5]: # Show first 5 errors
                        messages.error(request, err)
                        
                return redirect('product_list')

            except Exception as e:
                messages.error(request, f"Critical Error: {str(e)}")

    else:
        form = ImportFileForm()
    
    return render(request, 'inventory/product_import.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser)
def user_list(request):
    users = User.objects.all()
    return render(request, 'inventory/user_list.html', {'users': users})

@user_passes_test(lambda u: u.is_superuser)
def user_create(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'New user created and permissions assigned successfully.')
            return redirect('user_list')
    else:
        form = CustomUserCreationForm()
    return render(request, 'inventory/user_form.html', {'form': form, 'title': 'Create New User'})

@user_passes_test(lambda u: u.is_superuser)
def user_update(request, pk):
    user_to_edit = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=user_to_edit)
        if form.is_valid():
            form.save()
            messages.success(request, f'User {user_to_edit.username} updated successfully.')
            return redirect('user_list')
    else:
        form = CustomUserChangeForm(instance=user_to_edit)
    return render(request, 'inventory/user_form.html', {'form': form, 'title': f'Edit User: {user_to_edit.username}'})

@user_passes_test(lambda u: u.is_superuser)
def user_delete(request, pk):
    user_to_delete = get_object_or_404(User, pk=pk)
    
    # Prevent self-deletion
    if user_to_delete == request.user:
        messages.error(request, "You cannot delete your own account.")
        return redirect('user_list')
        
    if request.method == 'POST':
        user_to_delete.delete()
        messages.success(request, f"User {user_to_delete.username} deleted successfully.")
        return redirect('user_list')
    
    return render(request, 'inventory/user_confirm_delete.html', {'user_to_delete': user_to_delete})

# Test
# print("Before API views block") # Removed
# API Views
from rest_framework import viewsets
from .serializers import ProductSerializer, CategorySerializer
from .models import Category

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer

# Imports moved to the top

# print("Before pos_view") # Removed
@login_required
def pos_view(request):
    # Initialize cart in session if not present
    cart = request.session.get('cart', {})
    products = Product.objects.all()
    customers = Customer.objects.all() # Pass customers to template
    sales_people = SalesPerson.objects.filter(is_active=True)
    
    # Handle Cart Actions
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add':
            p_id = request.POST.get('product_id')
            # Assuming simple qty 1 add, or input
            try:
                qty = int(request.POST.get('quantity', 1))
            except ValueError:
                qty = 1
                
            product = get_object_or_404(Product, id=p_id)
            
            # Check physical stock (ignoring session cart for a moment, or sum it up)
            current_cart_qty = cart.get(str(p_id), {}).get('quantity', 0)
            if product.quantity < (current_cart_qty + qty):
                 messages.error(request, f"Not enough stock for {product.name}. Available: {product.quantity}")
            else:
                if str(p_id) in cart:
                    cart[str(p_id)]['quantity'] += qty
                else:
                    cart[str(p_id)] = {
                        'id': product.id,
                        'name': product.name,
                        'price': float(product.selling_price),
                        'quantity': qty
                    }
                request.session['cart'] = cart
                messages.success(request, f"Added {product.name}")

        elif action == 'remove':
            p_id = request.POST.get('product_id')
            if str(p_id) in cart:
                del cart[str(p_id)]
                request.session['cart'] = cart

        elif action == 'clear':
            request.session['cart'] = {}
            
        elif action == 'checkout':
            if not cart:
                messages.error(request, "Cart is empty")
            else:
                try:
                    # Get Customer
                    c_id = request.POST.get('customer_id')
                    customer = None
                    if not c_id:
                        messages.error(request, "Please select a Customer.")
                        return redirect('pos')
                        
                    customer = Customer.objects.filter(id=c_id).first()
                        
                    # Get Sales Person
                    sp_id = request.POST.get('sales_person_id')
                    sales_person = None
                    if not sp_id:
                        messages.error(request, "Please select a Sales Person.")
                        return redirect('pos')

                    sales_person = SalesPerson.objects.filter(id=sp_id).first()

                    # Create Sale
                    total_amt = sum(item['quantity'] * item['price'] for item in cart.values())
                    sale = Sale.objects.create(
                        user=request.user, 
                        total_amount=total_amt,
                        customer=customer,
                        sales_person=sales_person
                    )
                    
                    for p_id, item in cart.items():
                        product = Product.objects.get(id=p_id)
                        qty = item['quantity']
                        price = item['price']
                        
                        # Double check stock
                        if product.quantity < qty:
                            raise Exception(f"Stock changed! Not enough {product.name}")
                            
                        # Update Stock
                        product.quantity -= qty
                        product.save()
                        
                        # Create Item
                        SaleItem.objects.create(
                            sale=sale,
                            product=product,
                            quantity=qty,
                            unit_price=price 
                            # Note: models.py defines subtotal in save()
                        )
                        
                        # Log
                        StockLog.objects.create(
                            product=product,
                            user=request.user,
                            action='SALE',
                            quantity_change=-qty,
                            reason=f"Sale {sale.order_id}"
                        )
                        
                    # Success
                    request.session['cart'] = {}
                    messages.success(request, f"Sale {sale.order_id} completed successfully!")
                    return redirect('dashboard')

                except Exception as e:
                    messages.error(request, f"Error processing sale: {str(e)}")
                    # Delete sale if created? Transaction atomic needed ideally.
                    
    # Calculate Cart Total
    cart_total = sum(item['quantity'] * item['price'] for item in cart.values())
    
    return render(request, 'inventory/pos.html', {
        'products': products,
        'cart': cart,
        'cart_total': cart_total,
        'customers': customers,
        'sales_people': sales_people
    })

# Customer Views

@login_required
def customer_list(request):
    customers = Customer.objects.all()
    return render(request, 'inventory/customer_list.html', {'customers': customers})

@user_passes_test(lambda u: u.is_superuser)
def customer_import(request):
    if request.method == 'POST':
        form = ImportFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            created_count = 0
            errors = []
            
            try:
                # Identify file type
                is_excel = file.name.endswith(('.xlsx', '.xls'))
                
                rows = []
                if is_excel:
                    if not openpyxl:
                        messages.error(request, "Excel support not installed. Please use CSV.")
                        return redirect('customer_import')
                        
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    iter_rows = ws.iter_rows(values_only=True)
                    next(iter_rows, None) # Skip Header
                    rows = list(iter_rows)
                else:
                    # CSV handling
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.reader(decoded_file)
                    next(reader, None) # Skip Header
                    rows = list(reader)

                for idx, row in enumerate(rows, start=2):
                    try:
                        # Row: Name, Phone, Address, Email
                        if not row or len(row) < 2: 
                             continue 
                        
                        name = str(row[0]).strip()
                        phone = str(row[1]).strip()
                        address = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                        email = str(row[3]).strip() if len(row) > 3 and row[3] else None
                        
                        if not name or not phone:
                            errors.append(f"Row {idx}: Missing Name or Phone")
                            continue

                        if Customer.objects.filter(phone=phone).exists():
                            continue 
                        
                        Customer.objects.create(
                            name=name,
                            phone=phone,
                            address=address,
                            email=email
                        )
                        created_count += 1
                        
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                
                if created_count > 0:
                    messages.success(request, f"Successfully imported {created_count} customers.")
                else:
                    messages.warning(request, "No new records imported.")
                    
                if errors:
                    messages.error(request, f"Review errors: {'; '.join(errors[:3])}...")

                return redirect('customer_list')

            except Exception as e:
                messages.error(request, f"File Error: {str(e)}")

    else:
        form = ImportFileForm()
    
    return render(request, 'inventory/customer_import.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser)
def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f"Customer {customer.name} added successfully.")
            # If came from POS (next param), redirect back
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('customer_list')
    else:
        form = CustomerForm()
    return render(request, 'inventory/customer_form.html', {'form': form, 'title': 'Add New Customer'})

@user_passes_test(lambda u: u.is_superuser)
def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, "Customer updated.")
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'inventory/customer_form.html', {'form': form, 'title': 'Edit Customer'})

@user_passes_test(lambda u: u.is_superuser)
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer.delete()
        messages.success(request, "Customer deleted.")
        return redirect('customer_list')
    return render(request, 'inventory/customer_confirm_delete.html', {'customer': customer})

# Sales Person Views

@login_required
def salesperson_list(request):
    sales_people = SalesPerson.objects.all()
    return render(request, 'inventory/salesperson_list.html', {'sales_people': sales_people})

@user_passes_test(lambda u: u.is_superuser)
def salesperson_create(request):
    if request.method == 'POST':
        form = SalesPersonForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Sales Person added.")
            return redirect('salesperson_list')
    else:
        form = SalesPersonForm()
    return render(request, 'inventory/salesperson_form.html', {'form': form, 'title': 'Add Sales Person'})

@user_passes_test(lambda u: u.is_superuser)
def salesperson_update(request, pk):
    person = get_object_or_404(SalesPerson, pk=pk)
    if request.method == 'POST':
        form = SalesPersonForm(request.POST, instance=person)
        if form.is_valid():
            form.save()
            messages.success(request, "Sales Person updated.")
            return redirect('salesperson_list')
    else:
        form = SalesPersonForm(instance=person)
    return render(request, 'inventory/salesperson_form.html', {'form': form, 'title': 'Edit Sales Person'})

@user_passes_test(lambda u: u.is_superuser)
def salesperson_delete(request, pk):
    person = get_object_or_404(SalesPerson, pk=pk)
    if request.method == 'POST':
        person.delete()
        messages.success(request, "Sales Person deleted.")
        return redirect('salesperson_list')
    return render(request, 'inventory/salesperson_confirm_delete.html', {'person': person})

@user_passes_test(lambda u: u.is_superuser)
def salesperson_import(request):
    if request.method == 'POST':
        form = ImportFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            created_count = 0
            errors = []
            
            try:
                # Identify file type
                is_excel = file.name.endswith(('.xlsx', '.xls'))
                
                rows = []
                if is_excel:
                    if not openpyxl:
                        messages.error(request, "Excel support not installed. Please use CSV.")
                        return redirect('salesperson_import')
                        
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    iter_rows = ws.iter_rows(values_only=True)
                    next(iter_rows, None) # Skip Header
                    rows = list(iter_rows)
                else:
                    # CSV handling
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.reader(decoded_file)
                    next(reader, None) # Skip Header
                    rows = list(reader)

                for idx, row in enumerate(rows, start=2): # Start counting from row 2 (header is 1)
                    try:
                        # Row: Name, Phone, Email
                        if not row or len(row) < 2: 
                             continue 
                        
                        name = str(row[0]).strip()
                        phone = str(row[1]).strip()
                        email = str(row[2]).strip() if len(row) > 2 else None
                        
                        if not name or not phone:
                            errors.append(f"Row {idx}: Missing Name or Phone")
                            continue

                        # Create (Uniqueness check on Phone)
                        if SalesPerson.objects.filter(phone=phone).exists():
                            # errors.append(f"Row {idx}: Phone {phone} already exists (Skipped)")
                            continue # Skip silently or log? User asked for import, probably just skip duplicates.
                        
                        SalesPerson.objects.create(
                            name=name,
                            phone=phone,
                            email=email
                        )
                        created_count += 1
                        
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                
                if created_count > 0:
                    messages.success(request, f"Successfully imported {created_count} sales people.")
                else:
                    messages.warning(request, "No new records imported.")
                    
                if errors:
                    messages.error(request, f"Review errors: {'; '.join(errors[:3])}...")

                return redirect('salesperson_list')

            except Exception as e:
                messages.error(request, f"File Error: {str(e)}")

    else:
        form = ImportFileForm()
    
    return render(request, 'inventory/salesperson_import.html', {'form': form})

@login_required
@permission_required('inventory.add_product', raise_exception=True)
def purchase_view(request):
    products = Product.objects.all()
    if request.method == 'POST':
        p_id = request.POST.get('product_id')
        qty_str = request.POST.get('quantity')
        cost_str = request.POST.get('cost_price')
        supplier = request.POST.get('supplier')
        
        if not p_id:
            messages.error(request, "Please select a valid product from the dropdown.")
            return render(request, 'inventory/purchase_form.html', {'products': products})

        try:
            qty = int(qty_str)
            cost = float(cost_str)
        except (ValueError, TypeError):
             messages.error(request, "Invalid Quantity or Cost Price.")
             return render(request, 'inventory/purchase_form.html', {'products': products})
             
        product = get_object_or_404(Product, id=p_id)
        
        # Create Purchase Record
        purchase = Purchase.objects.create(
            user=request.user,
            supplier=supplier,
            total_amount=qty * cost
        )
        
        PurchaseItem.objects.create(
            purchase=purchase,
            product=product,
            quantity=qty,
            unit_cost=cost
        )
        
        # Update Product Stock
        product.quantity += qty
        product.purchase_price = cost # Update latest cost price? Optional but useful.
        product.save()
        
        # Log
        StockLog.objects.create(
            product=product,
            user=request.user,
            action='PURCHASE',
            quantity_change=qty,
            reason=f"Purchase {purchase.order_id}"
        )
        
        messages.success(request, f"Stock added for {product.name} (Ref: {purchase.order_id})")
        return redirect('product_list')
        
    return render(request, 'inventory/purchase_form.html', {'products': products})

@login_required
def report_view(request):
    # Simple Daily Sales Report
    sales = Sale.objects.order_by('-date')
    
    # Calculate Total Sales
    total_sales = sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Calculate Profit (Naive approach: SalePrice - CurrentPurchasePrice)
    # A better way is iterating items.
    
    daily_sales = Sale.objects.annotate(date_only=TruncDate('date')).values('date_only').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('-date_only')
    
    return render(request, 'inventory/reports.html', {
        'sales': sales,
        'total_sales': total_sales,
        'daily_sales': daily_sales
    })

@login_required
def transaction_detail(request, type, id):
    if type == 'sale':
        transaction = get_object_or_404(Sale, id=id)
        items = transaction.items.select_related('product').all()
        # For sales, we show selling price
        # item.unit_price is stored
    elif type == 'purchase':
        transaction = get_object_or_404(Purchase, id=id)
        items = transaction.items.select_related('product').all()
    else:
        return redirect('dashboard')
        
    return render(request, 'inventory/transaction_detail.html', {
        'type': type,
        't': transaction,
        'items': items
    })

@login_required
def salesperson_import(request):
    if request.method == 'POST':
        form = ImportFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            created_count = 0
            errors = []
            
            try:
                # Identify file type
                is_excel = file.name.endswith(('.xlsx', '.xls'))
                
                rows = []
                if is_excel:
                    if not openpyxl:
                        messages.error(request, "Excel support not installed. Please use CSV.")
                        return redirect('salesperson_import')
                        
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    iter_rows = ws.iter_rows(values_only=True)
                    next(iter_rows, None) # Skip Header
                    rows = list(iter_rows)
                else:
                    # CSV handling
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.reader(decoded_file)
                    next(reader, None) # Skip Header
                    rows = list(reader)

                for idx, row in enumerate(rows, start=2): # Start counting from row 2 (header is 1)
                    try:
                        # Row: Name, Phone, Email
                        if not row or len(row) < 2: 
                             continue 
                        
                        name = str(row[0]).strip()
                        phone = str(row[1]).strip()
                        email = str(row[2]).strip() if len(row) > 2 else None
                        
                        if not name or not phone:
                            errors.append(f"Row {idx}: Missing Name or Phone")
                            continue

                        # Create (Uniqueness check on Phone)
                        if SalesPerson.objects.filter(phone=phone).exists():
                            # errors.append(f"Row {idx}: Phone {phone} already exists (Skipped)")
                            continue # Skip silently or log? User asked for import, probably just skip duplicates.
                        
                        SalesPerson.objects.create(
                            name=name,
                            phone=phone,
                            email=email
                        )
                        created_count += 1
                        
                    except Exception as e:
                        errors.append(f"Row {idx}: {str(e)}")
                
                if created_count > 0:
                    messages.success(request, f"Successfully imported {created_count} sales people.")
                else:
                    messages.warning(request, "No new records imported.")
                    
                if errors:
                    messages.error(request, f"Review errors: {'; '.join(errors[:3])}...")

                return redirect('salesperson_list')

            except Exception as e:
                messages.error(request, f"File Error: {str(e)}")

    else:
        form = ImportFileForm()
    
    return render(request, 'inventory/salesperson_import.html', {'form': form})
